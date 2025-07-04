from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import List, Optional
from .models import Task
from uuid import UUID
from datetime import datetime

EXCEL_FILE = Path(__file__).parent / "tasks.xlsx"

HEADERS = [
    "asigner", "asignee", "request_date", "task_name", "task_description", "deadline", "status", "task_id"
]

def init_excel():
    if not EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

def add_task(task: Task):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        task.asigner,
        task.asignee,
        task.request_date.isoformat(),
        task.task_name,
        task.task_description,
        task.deadline.isoformat() if task.deadline else None,
        task.status,
        str(task.task_id)
    ])
    wb.save(EXCEL_FILE)

def list_tasks(asignee: Optional[str] = None) -> List[Task]:
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Handle both old and new schema rows
        if not row:
            continue
        # Old schema: 7 columns (no task_name, task_description)
        if len(row) == 7:
            asigner, asignee, request_date, task_legacy, deadline, status, task_id = row
            task_name = task_legacy or "(no name)"
            task_description = task_legacy or "(no description)"
        # New schema: 8 columns
        elif len(row) == 8:
            asigner, asignee, request_date, task_name, task_description, deadline, status, task_id = row
        else:
            print(f"Skipping row with unexpected column count: {row}")
            continue
        # Skip if required fields are missing
        if any(x is None or str(x).strip() == "" for x in [asigner, asignee, request_date, task_name, task_description, deadline, status, task_id]):
            continue
        if asignee and asignee != asignee:
            continue
        try:
            task = Task(
                asigner=asigner,
                asignee=asignee,
                request_date=datetime.fromisoformat(request_date),
                task_name=task_name,
                task_description=task_description,
                deadline=datetime.fromisoformat(deadline) if deadline else None,
                status=status,
                task_id=UUID(task_id)
            )
            tasks.append(task)
        except Exception as e:
            print(f"Skipping invalid row in Excel: {row} (error: {e})")
            continue
    return tasks

def mark_task_complete(task_id: UUID, status: str = "complete") -> bool:
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[6].value) == str(task_id):
            row[5].value = status
            wb.save(EXCEL_FILE)
            return True
    return False

def update_task_status_by_name(asignee: str, task_name: str, status: str) -> bool:
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        # row[1] = asignee, row[3] = task
        if row[1].value == asignee and row[3].value == task_name:
            row[5].value = status
            wb.save(EXCEL_FILE)
            return True
    return False

# Update all fields of a task by its unique task_id
def update_task_by_id(task_id: UUID, updated_task: Task) -> bool:
    """
    Update all fields of a task in the Excel file by its task_id.
    Returns True if the update was successful, False if not found.
    """
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        # The task_id is always in the last column (index 7)
        if str(row[7].value) == str(task_id):
            row[0].value = updated_task.asigner
            row[1].value = updated_task.asignee
            row[2].value = updated_task.request_date.isoformat()
            row[3].value = updated_task.task_name
            row[4].value = updated_task.task_description
            row[5].value = updated_task.deadline.isoformat() if updated_task.deadline else None
            row[6].value = updated_task.status
            row[7].value = str(updated_task.task_id)
            wb.save(EXCEL_FILE)
            return True
    return False
